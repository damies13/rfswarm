import os

from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from pandas import read_excel
from robot.api import logger as LOGGER


def read_all_xlsx_sheets(xlsx_path: str) -> list:
	return load_workbook(xlsx_path, keep_links=False).sheetnames


def read_all_xlsx_text_data(xlsx_path: str) -> list:
	"""
	Read all xlsx text data from all sheets.
	"""
	wb = load_workbook(xlsx_path, keep_links=False)

	xlsx_tables = []
	for sheet in wb.sheetnames:
		sheet_content = read_xlsx_text_data_from_sheet(xlsx_path, sheet)
		xlsx_tables.append(sheet_content)

	print(xlsx_tables)
	return xlsx_tables


def get_xlsx_sheet_by_name(xlsx_path: str, sheet_name: str):
	"""
	Find the full name of the xlsx sheet by its name.
	For example: given: "Contents tables" output: "4 Contents tables".
	"""
	xlsx_sheets = load_workbook(xlsx_path, keep_links=False).sheetnames
	for sheet in xlsx_sheets:
		if " " in list(sheet):
			sheet_fragmented = sheet.split(" ", maxsplit=1)
			if sheet_fragmented[1] == sheet_name:
				return sheet

	return 0


def read_xlsx_text_data_from_sheet(xlsx_path: str, sheet: str, start_at=None, stop_at=None) -> list:
	"""
	Read xlsx text data from the given xlsx sheet.
	start_at - start from the specified string, but do not include it.
	stop_at - end at specified string, and don't include it.
	"""
	all_excel_data_frame = read_excel(xlsx_path, sheet, keep_default_na=False, na_values=['NaN'])
	excel_data_list = all_excel_data_frame.values.tolist()

	for row_n in range(len(excel_data_list) - 1, -1, -1):
		for cell_n in range(len(excel_data_list[row_n]) - 1, -1, -1):
			if excel_data_list[row_n][cell_n] == "":
				excel_data_list[row_n].pop(cell_n)

		if len(excel_data_list[row_n]) == 0:
			excel_data_list.pop(row_n)

	start_index = 0
	end_index = len(excel_data_list)
	if start_at is not None:
		for row_n in range(0, len(excel_data_list)):
			if start_at in excel_data_list[row_n]:
				start_index = row_n + 1
	if stop_at is not None:
		for row_n in range(0, len(excel_data_list)):
			if stop_at in excel_data_list[row_n]:
				end_index = row_n

	return excel_data_list[start_index: end_index]


def extract_image_from_xlsx_sheet(xlsx_path: str, sheet: str, cell_id: str, output_folder: str, show_image: bool = False):
	"""
	Extract an image from XLSX file from a given cell in specified sheet.
	Returns name of the saved image.
	"""
	wb = load_workbook(xlsx_path, keep_links=False)
	sheet_obj = wb[sheet]
	try:
		image_loader = SheetImageLoader(sheet_obj)
	except Exception as e:
		raise AssertionError("Error in SheetImageLoader:", e)
	try:
		image = image_loader.get(cell_id)
	except Exception as e:
		print(f"Cell {cell_id} doesn't contain an image. Exception: {e}")
		return 0
	if show_image:
		image.show()
	img_name = (sheet + "_" + cell_id + "_image.png").replace(" ", "_")

	if not os.path.exists(output_folder):
		os.makedirs(output_folder)

	image.save(os.path.join(output_folder, img_name))
	print(f"XLSX RFSwarm raport image saved as: {img_name}")

	return img_name


def get_font_name_from_xlsx_sheet(xlsx_path: str, sheet: str) -> str:
	"""Returns the name of the font that is used by default in given sheet."""
	wb = load_workbook(xlsx_path, keep_links=False)
	sheet_obj = wb[sheet]

	font_name = ''
	for row in sheet_obj.iter_rows():
		for cell in row:
			if cell.font and cell.value is not None:
				if cell.font.name != font_name and font_name != '':
					LOGGER.warn(f'Different fonts found in the sheet! font:{font_name}')
				font_name = cell.font.name

	return font_name
